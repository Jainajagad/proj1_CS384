import os
import csv
import openpyxl
import shutil
from openpyxl.styles import colors
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
import yagmail

def generate_marksheet(path1, path2, pos, neg):

	pos = int(pos)
	neg = int(neg)

	if not os.path.isdir("sample_output"):
		os.mkdir("sample_output")

	if not os.path.isdir(os.path.join("sample_output", "marksheet")):
		os.mkdir(os.path.join("sample_output", "marksheet"))

	if path1.split('\\')[-1] != "master_roll.csv":
		if os.path.isfile(path1):
			os.remove(path1)

	with open(path1, "r") as csvfile:
		reader = csv.reader(csvfile, delimiter=',', skipinitialspace=True)
		roll_name_map= {}
		for row in reader:
			roll_name_map[row[0]] = row[1]
	# print(roll_name_map)

	with open(path2, "r") as csvfile:

		global stud_info
		stud_info = {}
		reader = csv.reader(csvfile, delimiter=',', skipinitialspace=True)
		header = next(reader)
		f_header = []
		f_header.extend(header[1:5])
		f_header.append(header[6])
		f_header.append("Options")

		# print(f_header)
		for row in reader:
			# print(type(row[6]))
			stud_info[row[6]] = {}
			i=1
			for h in f_header:
				if i==4:
					break
				stud_info[row[6]][h] = row[i]
				i+=1

			stud_info[row[6]][f_header[4]] = row[6]
			stud_info[row[6]][f_header[5]] = row[7:]

	# print(stud_info)
	if "ANSWER" not in stud_info.keys():
		if path2.split('\\')[-1] != "responses.csv":
			if os.path.isfile(path2):
				os.remove(path2)
		return r"no roll number with ANSWER is present, Cannot Process!"

	# print(stud_info)

	f_stud_data = {}

	correct_answers = stud_info["ANSWER"]["Options"]
	total_marks = len(correct_answers)*pos

	# print(pos, neg)

	for roll in stud_info.keys():
		if roll != "ANSWER":

			correct = 0
			missing = 0
			incorrect =0

			f_stud_data[roll] = {}
			f_stud_data[roll]["Name"] = stud_info[roll]["Name"]
			f_stud_data[roll]["Roll Number"] = stud_info[roll]["Roll Number"]

			for idx,val in enumerate(stud_info[roll]["Options"]):
				if val == correct_answers[idx]:
					correct += 1
				elif val != correct_answers[idx]:
					if val == "":
						missing += 1
					else:
						incorrect += 1

			marks_stud = (correct* pos) + (incorrect* neg)

			f_stud_data[roll]["right_no"] = correct
			f_stud_data[roll]["wrong_no"] = incorrect
			f_stud_data[roll]["not_attempt"] = missing
			f_stud_data[roll]["max_no"] = correct + incorrect + missing
			f_stud_data[roll]["total_right"] = correct* pos
			f_stud_data[roll ]["total_wrong"] = incorrect* neg
			f_stud_data[roll]["total_max"] = str(marks_stud) + "/" + str(total_marks)
	

	for roll in f_stud_data.keys():

		wb = Workbook()

		sheet = wb.active

		row_count = 60
		column_count = 5

		for i in range(1, row_count + 1):
			for j in range(1, column_count + 1):
				sheet.column_dimensions[get_column_letter(j)].width = 17
				# print(i,j)
				sheet.cell(row=i, column=j).font = Font(name= "Century", size=12)

		sheet.title = "quiz"

		img = openpyxl.drawing.image.Image('IITP_Logo.jpg')
		img.anchor = 'A1'
		sheet.add_image(img)
		sheet.merge_cells('A5:E5')

		sheet["A5"].value = 'Mark Sheet'
		sheet["A5"].font = Font(name="Century",size=18, bold=True, underline = "single")
		sheet["A5"].border = Border()
		sheet["A5"].alignment = Alignment(horizontal="center", vertical="bottom")

		sheet["A6"].value = "Name:"
		sheet["A6"].alignment = Alignment(horizontal="right")
		sheet["B6"].value = f_stud_data[roll]["Name"]
		sheet["B6"].font = Font(name= "Century", size=12, bold=True)
		sheet["B6"].alignment = Alignment(horizontal= "left")
		sheet["D6"].value = "Exam:"
		sheet["D6"].alignment = Alignment(horizontal="right")
		sheet["E6"].value = "quiz"
		sheet["E6"].font = Font(name= "Century", size=12, bold=True)
		sheet["E6"].alignment = Alignment(horizontal= "left")

		sheet["A7"].value = "Roll Number:"
		sheet["A7"].alignment = Alignment(horizontal="right")
		sheet["B7"].value = f_stud_data[roll]["Roll Number"]
		sheet["B7"].font = Font(name= "Century", size=12, bold=True)
		sheet["B7"].alignment = Alignment(horizontal= "left")

		sheet.append(["","","","",""])

		marks_info = []
		marks_info.append(tuple(["", "Right", "Wrong", "Not Attempt", "Max"]))
		marks_info.append(tuple(["No.", f_stud_data[roll]["right_no"], f_stud_data[roll]["wrong_no"],
		                    f_stud_data[roll]["not_attempt"], f_stud_data[roll]["max_no"]]))
		marks_info.append(tuple(["Marking", pos, neg, 0, ""]))
		marks_info.append(tuple(["Total", f_stud_data[roll]["total_right"], f_stud_data[roll]["total_wrong"],
			               "", f_stud_data[roll]["total_max"]]))

		marks_info = tuple(marks_info)
		thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

		i1=0
		for i in range(9, 13):
			j1=0
			for j in range(1, 6):
				sheet.cell(row=i, column=j).value = marks_info[i1][j1]
				sheet.cell(row=i, column=j).border = thin_border
				sheet.cell(row=i, column=j).alignment = Alignment(horizontal= "center")

				if i1==0 or (j1 == 0 and i1>0):
					sheet.cell(row=i, column=j).font = Font(bold = True, name= "Century", size=12)

				if i1 > 0:
					if j1 == 1:
						sheet.cell(row=i, column=j).font = Font(color = "00008000", name= "Century", size=12)
					elif j1 == 2:
						sheet.cell(row=i, column=j).font = Font(color = "00FF0000", name= "Century", size=12)

				if i1 == 3 and j1 == 4:
					sheet.cell(row=i, column=j).font = Font(color = "000000FF", name= "Century", size=12)
				j1+=1
			i1+=1

		sheet.append(["","","","",""])
		sheet.append(["","","","",""])

		s_ans = stud_info[roll]["Options"]
		c_ans = stud_info["ANSWER"]["Options"]
		c1 = c_ans.copy()

		# print("c",c_ans)

		turn =1

		while 1:
			# print(turn)
			if turn ==7:
				# print("****")
				break

			stud_ans = []

			stud_ans.append(tuple(["Student Ans", "Correct Ans"]))
			if len(s_ans) < 25:
				f_s_ans = s_ans[:len(s_ans)]
			elif len(s_ans) >= 25:
				f_s_ans = s_ans[:25]

			if len(c1) < 25:
				f_c_ans = c1[:len(c1)]
			elif len(s_ans) >= 25:
				f_c_ans = c1[:25]
			# print(s_ans)
			# print(c_ans)

			for k in range(len(f_s_ans)):
				s_ans.pop(0)

			for k1 in range(len(f_c_ans)):
				c1.pop(0)

			for x in range(len(f_c_ans)):
				stud_ans.append(tuple([f_s_ans[x], f_c_ans[x]]))

			stud_ans = tuple(stud_ans)
			# print(stud_ans)

			end = 15+ len(stud_ans)

			thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

			i1=0 
			for i in range(15, end):
				j1=0
				for j in range(turn, turn+2):
					sheet.cell(row=i, column=j).value =stud_ans[i1][j1]
					sheet.cell(row=i, column=j).border = thin_border
					sheet.cell(row=i, column=j).alignment = Alignment(horizontal= "center")

					if i1==0:
						sheet.cell(row=i, column=j).font = Font(bold = True, name= "Century", size=12)

					if i1>0:
						if (stud_ans[i1][0] == stud_ans[i1][1]):
							sheet.cell(row=i, column=j).font = Font(color = "00008000", name= "Century", size=12)
						else:
							if stud_ans[i1][0] != '':
								sheet.cell(row=i, column=j).font = Font(color = "00FF0000", name= "Century", size=12)

					if i1>0 and j1==1:
						sheet.cell(row=i, column=j).font = Font(color = "000000FF", name= "Century", size=12)
					j1+=1
				i1+=1

			turn+=3

		name = os.path.join(os.getcwd(), "sample_output", "marksheet",(roll+".xlsx"))
		wb.save(name)

def concise_marksheet(path1, path2, pos, neg):

	pos = int(pos)
	neg = int(neg)

	if path1.split('\\')[-1] != "master_roll.csv":
		if os.path.isfile(path1):
			os.remove(path1)

	with open(path2, "r") as csvfile:

		stud_info = {}
		reader = csv.reader(csvfile, delimiter=',', skipinitialspace=True)
		header = next(reader)
		# print(header)
		f_header = []
		f_header.extend(header[0:7])
		f_header.append("Options")

		# print(f_header)

		for row in reader:
			# print(type(row[6]))
			stud_info[row[6]] = {}
			i=0
			for h in f_header:
				if i==7:
					break
				stud_info[row[6]][h] = row[i]
				i+=1

			stud_info[row[6]][f_header[7]] = row[7:]

	if "ANSWER" not in stud_info.keys():
		if path2.split('\\')[-1] != "responses.csv":
			if os.path.isfile(path2):
				os.remove(path2)
		return r"no roll number with ANSWER is present, Cannot Process!"

	# print(stud_info)
	correct_answers = stud_info["ANSWER"]["Options"]
	total_marks = len(correct_answers)*pos

	# print(total_marks)
	# print(pos, neg)
	f_stud_data = {}

	for roll in stud_info.keys():

		correct = 0
		missing = 0
		incorrect =0

		f_stud_data[roll] = {}
		f_stud_data[roll]["Name"] = stud_info[roll]["Name"]
		f_stud_data[roll]["Roll Number"] = stud_info[roll]["Roll Number"]

		for idx,val in enumerate(stud_info[roll]["Options"]):
			if val == correct_answers[idx]:
				correct += 1
			elif val != correct_answers[idx]:
				if val == "":
					missing += 1
				else:
					incorrect += 1


		marks_stud = (correct* pos) + (incorrect* neg)
		# print(marks_stud, total_marks)
		f_stud_data[roll]["status_Ans"] = str([correct, incorrect, missing])
		f_stud_data[roll]["score_after_neg"] = str(marks_stud) + "/" + str(total_marks)

	concise_info = []
	header =["Timestamp", "Email address", "Google Score", "Name", "IITP webmail", "Phone (10 digit only)", "Score_After_Negative", "Roll Number"]
	end = 7+ (len(correct_answers)-1)

	for i in range(7, end+1):
		header.append("Unnamed: "+ str(i))
	header.append("statusAns")

	for roll in stud_info.keys():
		s_inf = []
		s_inf.append(stud_info[roll]["Timestamp"])
		s_inf.append(stud_info[roll]["Email address"])
		s_inf.append(stud_info[roll]["Score"])
		s_inf.append(stud_info[roll]["Name"])
		s_inf.append(stud_info[roll]["IITP webmail"])
		s_inf.append(stud_info[roll]["Phone (10 digit only)"])
		s_inf.append(f_stud_data[roll]["score_after_neg"])
		s_inf.append(stud_info[roll]["Roll Number"])
		s_inf.extend(stud_info[roll]["Options"])
		s_inf.append(f_stud_data[roll]["status_Ans"])
		concise_info.append(s_inf)

	if not os.path.isdir("sample_output"):
		os.mkdir("sample_output")

	if not os.path.isdir("sample_output//marksheet"):
		os.mkdir("sample_output//marksheet")

	with open(os.path.join("sample_output", "marksheet","concise_marksheet.csv"), "w", newline='') as f:
		csvwriter = csv.writer(f)
		csvwriter.writerow(header)
		csvwriter.writerows(concise_info)

def Send_email(path1, path2, pos, neg):

	if path1.split('\\')[-1] != "master_roll.csv":
		if os.path.isfile(path1):
			os.remove(path1)

	with open(path2, "r") as csvfile:

		stud_info = {}
		reader = csv.reader(csvfile, delimiter=',', skipinitialspace=True)
		header = next(reader)
		# print(header)
		f_header = []
		f_header.extend(header[0:7])
		f_header.append("Options")

		# print(f_header)

		for row in reader:
			# print(type(row[6]))
			stud_info[row[6]] = {}
			i=0
			for h in f_header:
				if i==7:
					break
				stud_info[row[6]][h] = row[i]
				i+=1

			stud_info[row[6]][f_header[7]] = row[7:]

	print(stud_info)

	if "ANSWER" not in stud_info.keys():
		if path2.split('\\')[-1] != "responses.csv":
			if os.path.isfile(path2):
				os.remove(path2)
		return r"no roll number with ANSWER is present, Cannot Process!"

	'''
	for roll in stud_info.keys():
		filename = os.path.join("sample_output", "marksheet", (roll+ ".xlsx"))
	'''